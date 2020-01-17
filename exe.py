from selenium import webdriver
import xlsxwriter   
from openpyxl import Workbook
from openpyxl import load_workbook

path = r"E:\ml\Web_Scraping\chromedriver"
wb = load_workbook("Results_CSE.xlsx")
ws = wb.worksheets[0] 
usn=''

for num in range(142,191):
    try:
        driver = webdriver.Chrome(executable_path = path)
        driver.get('http://results.jssstuniv.in/')
        
        
        workbook = xlsxwriter.Workbook('Results_CSE.xlsx') 
        worksheet = workbook.add_worksheet()  
        
        cand_result=[]
        
        inputElement = driver.find_element_by_id("USN")
        usn='01jst17cs'+format(num,'03d')
        inputElement.send_keys(usn)
        driver.find_element_by_class_name('button2').click()
         
        web_name= driver.find_element_by_tag_name('h1')
        name=web_name.text
        cand_result.append(name)
           
        web_grade1=driver.find_elements_by_id('grade1')
        for i in web_grade1:
            grade1=i.text
        cand_result.append(grade1)
            
        web_grade2=driver.find_elements_by_id('grade2')
        for i in web_grade2:
            grade2=i.text
        cand_result.append(grade2)
            
        web_grade3=driver.find_elements_by_id('grade3')
        for i in web_grade3:
            grade3=i.text
        cand_result.append(grade3)
            
        web_grade4=driver.find_elements_by_id('grade4')
        for i in web_grade4:
            grade4=i.text
        cand_result.append(grade4)
            
        web_grade5=driver.find_elements_by_id('grade5')
        for i in web_grade5:
            grade5=i.text
        cand_result.append(grade5)
            
        web_grade6=driver.find_elements_by_id('grade6')
        for i in web_grade6:
            grade6=i.text
        cand_result.append(grade6)
        flag=1  
        while(flag):    
            inputElement = driver.find_element_by_id("cred1")
            inputElement.send_keys('4')
            
            inputElement = driver.find_element_by_id("cred2")
            inputElement.send_keys('5')
            
            inputElement = driver.find_element_by_id("cred3")
            inputElement.send_keys('5')
            
            inputElement = driver.find_element_by_id("cred4")
            inputElement.send_keys('4')
            
            inputElement = driver.find_element_by_id("cred5")
            inputElement.send_keys('5')
            
            inputElement = driver.find_element_by_id("cred6")
            inputElement.send_keys('5')
            
            
            
            driver.find_element_by_id("sgpa").click()
            web_sgpa=driver.find_elements_by_id('sgpa')
            for i in web_sgpa:
                sgpa=i.text[7:]
            if(sgpa!='te SGPA'):
                flag=0
                
        cand_result.append(sgpa)
        driver.find_element_by_class_name('button2').click()
        
        
        ws.append(cand_result)
        wb.save("Results_CSE.xlsx")

        driver.close()
    except:
        print(usn)  
        driver.close()
     